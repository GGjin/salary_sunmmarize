from openpyxl import load_workbook
from decimal import Decimal
import re

wb = load_workbook("aa.xlsx", data_only=True)
sheet_names = wb.sheetnames
sheet_1 = wb[sheet_names[0]]

salary_data = []

for row in range(3, sheet_1.max_row + 1):
    id = sheet_1.cell(row, 2).value
    if id is not None:
        a = {"id": id, "1月": None, "2月": None, "3月": None, "4月": None, "5月": None,
             "6月": None, "7月": None, "8月": None, "9月": None, "10月": None, "11月": None,
             "12月": None}
        salary_data.append(a)

# print(salary_data)
for title in sheet_names[1:-1:]:
    sheet = wb[title]
    for row in range(3, sheet.max_row + 1):
        id = str(sheet.cell(row, 35).value)
        for data in salary_data:
            if id == data["id"]:
                salary = sheet.cell(row, 31).value
                data[title] = salary
print(salary_data)
month = {1: "1月", 2: "2月", 3: "3月", 4: "4月", 5: "5月", 6: "6月", 7: "7月", 8: "8月", 9: "9月", 10: "10月", 11: "11月",
         12: "12月", }
# for row in sheet_1["B3:P15"]:
#     print(row[1+1].value)
for row in sheet_1["B3:P15"]:
    for data in salary_data:
        if row[0].value == data["id"]:
            for key in month.keys():
                row[key + 2].value = data[month[key]]
#
wb.save("cc.xlsx")
print("----完成----")
